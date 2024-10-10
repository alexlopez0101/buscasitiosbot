import os
import telebot
import openpyxl
from dotenv import load_dotenv
import requests
from io import BytesIO
from flask import Flask, request
import logging

# Configurar logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Cargar variables de entorno
load_dotenv()

# Configurar el bot
BOT_TOKEN = os.getenv('BOT_TOKEN')
if not BOT_TOKEN:
    raise ValueError("No se encontró el token del bot. Asegúrate de tener un archivo .env con BOT_TOKEN=tu_token")

bot = telebot.TeleBot(BOT_TOKEN)

# Configurar Flask
app = Flask(__name__)

# Cargar el archivo Excel desde Google Drive
EXCEL_URL = os.getenv('EXCEL_URL')
if not EXCEL_URL:
    raise ValueError("No se encontró la URL del Excel. Asegúrate de tener un archivo .env con EXCEL_URL=tu_url")

try:
    logger.debug(f"Intentando descargar el archivo Excel desde: {EXCEL_URL}")
    response = requests.get(EXCEL_URL, allow_redirects=True)
    response.raise_for_status()  # Esto lanzará una excepción para códigos de estado HTTP no exitosos
    
    logger.debug(f"Archivo descargado exitosamente. Tamaño: {len(response.content)} bytes")
    logger.debug(f"Tipo de contenido: {response.headers.get('Content-Type')}")
    
    excel_content = BytesIO(response.content)
    workbook = openpyxl.load_workbook(excel_content, data_only=True)
    logger.info("Archivo Excel cargado exitosamente desde Google Drive")
    
    # Imprimir información sobre las hojas del libro
    logger.debug(f"Hojas en el libro: {workbook.sheetnames}")
    for sheet in workbook:
        logger.debug(f"Hoja: {sheet.title}, Filas: {sheet.max_row}, Columnas: {sheet.max_column}")

except requests.RequestException as e:
    logger.error(f"Error al descargar el archivo Excel: {e}")
    logger.debug(f"Respuesta de la solicitud: {e.response.text if e.response else 'No hay respuesta'}")
    exit(1)
except Exception as e:
    logger.error(f"Error inesperado al cargar el archivo Excel: {e}")
    exit(1)

# El resto de tu código (funciones buscar_por_id, buscar_por_nombre, etc.) va aquí...

@app.route('/' + BOT_TOKEN, methods=['POST'])
def getMessage():
    bot.process_new_updates([telebot.types.Update.de_json(request.stream.read().decode("utf-8"))])
    return "!", 200

@app.route("/")
def webhook():
    bot.remove_webhook()
    bot.set_webhook(url='https://your-app-name.herokuapp.com/' + BOT_TOKEN)
    return "!", 200

if __name__ == "__main__":
    logger.info("Iniciando la aplicación Flask")
    app.run(host="0.0.0.0", port=int(os.environ.get('PORT', 5000)))